import openpyxl
from openpyxl.styles import PatternFill, Font

GREEN='FFD5F5E3'; RED='FFFADBD8'; NOFILL=PatternFill(fill_type=None)
wb = openpyxl.load_workbook('PG_Commercial_Comparison_July2026.xlsx')

def numeric(v):
    if isinstance(v,(int,float)): return v
    return None

def recolor_row(ws, row, cols, clear_first=True):
    vals = {c: numeric(ws.cell(row=row,column=c).value) for c in cols}
    nums = {c:v for c,v in vals.items() if v is not None}
    if clear_first:
        for c in cols:
            ws.cell(row=row,column=c).fill = NOFILL
    if len(nums) < 2:
        return
    mn = min(nums.values()); mx = max(nums.values())
    if mn == mx:
        return
    for c,v in nums.items():
        if v == mn:
            ws.cell(row=row,column=c).fill = PatternFill('solid', fgColor=GREEN)
        elif v == mx:
            ws.cell(row=row,column=c).fill = PatternFill('solid', fgColor=RED)

# Instrument Rate Matrix: cols C=3 Cashfree, D=4 PayU, E=5 EaseBuzz, F=6 Razorpay
ws = wb['Instrument Rate Matrix']
for r in list(range(7,13)) + list(range(14,20)) + list(range(21,24)) + list(range(25,32)) + list(range(33,38)) + [39,40]:
    recolor_row(ws, r, [3,4,5,6])
ws.cell(row=3, column=1, value='Rates = MDR % excl. GST  |  Green = cheapest at this instrument  |  Red = most expensive at this instrument  |  ⚠ = key differentiator  |  * = rate assumed/interpolated  |  Italic = not re-quoted, April rate retained')

# Rate Card Reference: cols B=2 Cashfree, C=3 PayU, D=4 EaseBuzz, E=5 Razorpay
ws2 = wb['Rate Card Reference']
for r in [7,8,9,10,11,12,14,15,16,18,19,20,21,23,24,25,27,28,29,30,31,33,34]:
    recolor_row(ws2, r, [2,3,4,5])
ws2.cell(row=4, column=1, value='All rates are % of transaction value unless stated otherwise. Green = cheapest at this instrument, Red = most expensive. Italic = not re-quoted, April rate retained.')

# Cost by Payment Mode: cols E=5 Cashfree, F=6 PayU, G=7 EaseBuzz, H=8 Razorpay
ws3 = wb['Cost by Payment Mode']
for r in range(6,19):
    recolor_row(ws3, r, [5,6,7,8])
ws3.cell(row=4, column=1, value='All costs include 18% GST | Cashfree unchanged (no new quote) | Green = cheapest PG for that mode | Red = most expensive PG for that mode')

# Now mark "not re-quoted / carried forward" cells with italic font instead of fill, using the
# same carry-forward maps used when building this version, so it doesn't collide with cheapest/most-expensive colors.
carried_irm = {  # Instrument Rate Matrix: row -> list of cols carried forward unchanged
 9:[6],10:[5,6],11:[5],12:[4,5,6],15:[5,6],16:[3,4,5,6],17:[4,5],18:[3,4,5,6],19:[5,6],
 29:[5],30:[4,5,6],35:[4,5],39:[4,5,6],40:[4,5,6],
}
for r,cols in carried_irm.items():
    for c in cols:
        cell = ws.cell(row=r,column=c)
        cell.font = Font(italic=True, color='FF666666')

carried_rcr = {
 9:[3,4,5],10:[4,5],11:[3,4,5],12:[4,5],15:[4],16:[3,4],19:[5],20:[4,5],21:[4,5],
 25:[4],29:[3,4],33:[3,4,5],34:[3,4,5],
}
for r,cols in carried_rcr.items():
    for c in cols:
        cell = ws2.cell(row=r,column=c)
        cell.font = Font(italic=True, color='FF666666')

wb.save('PG_Commercial_Comparison_July2026.xlsx')
print("Colors fixed")
