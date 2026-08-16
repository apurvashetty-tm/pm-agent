import openpyxl
from openpyxl.styles import PatternFill, Font

GREEN='FFD5F5E3'; RED='FFFADBD8'; WHITE='FFFFFFFF'; LIGHTBLUE='FFEBF5FB'

wb = openpyxl.load_workbook('PG_Commercial_Comparison_July2026.xlsx')

# --- Summary Dashboard: Monthly Cost Comparison table, rows 11-14, col D = Total Cost ---
ws = wb['Summary Dashboard']
rows = [11,12,13,14]  # Cashfree, PayU, EaseBuzz, Razorpay
totals = {r: ws.cell(row=r, column=4).value for r in rows}
mn_row = min(totals, key=lambda r: totals[r])
mx_row = max(totals, key=lambda r: totals[r])
neutral_cycle = [WHITE, LIGHTBLUE, WHITE, LIGHTBLUE]
for i, r in enumerate(rows):
    if r == mn_row:
        fill = PatternFill('solid', fgColor=GREEN)
    elif r == mx_row:
        fill = PatternFill('solid', fgColor=RED)
    else:
        fill = PatternFill('solid', fgColor=neutral_cycle[i])
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = fill
print("Summary Dashboard fix: cheapest row =", ws.cell(row=mn_row,column=1).value, "| costliest row =", ws.cell(row=mx_row,column=1).value)

# --- Juspay Add-on: rows 22-25, col D = Combined Total. Row 25 (Razorpay) is "not compatible", not a price ---
ws2 = wb['Juspay Add-on']
compat_rows = [22,23,24]  # Cashfree+Juspay, PayU+Juspay, EaseBuzz+Juspay
totals2 = {r: ws2.cell(row=r, column=4).value for r in compat_rows}
mn2 = min(totals2, key=lambda r: totals2[r])
mx2 = max(totals2, key=lambda r: totals2[r])
neutral2 = {22:WHITE, 23:LIGHTBLUE, 24:WHITE}
for r in compat_rows:
    if r == mn2:
        fill = PatternFill('solid', fgColor=GREEN)
    elif r == mx2:
        fill = PatternFill('solid', fgColor=RED)
    else:
        fill = PatternFill('solid', fgColor=neutral2[r])
    for c in range(1, 6):
        ws2.cell(row=r, column=c).fill = fill
# Row 25 Razorpay: not a price comparison, it's "incompatible" — use neutral grey fill + italic, not red
for c in range(1, 6):
    cell = ws2.cell(row=25, column=c)
    cell.fill = PatternFill('solid', fgColor='FFF2F3F4')
    cell.font = Font(italic=True, color='FF666666')
print("Juspay Add-on fix: cheapest combo =", ws2.cell(row=mn2,column=1).value, "| costliest compatible combo =", ws2.cell(row=mx2,column=1).value)

wb.save('PG_Commercial_Comparison_July2026.xlsx')
