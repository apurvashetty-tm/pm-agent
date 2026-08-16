import openpyxl
from openpyxl.styles import PatternFill, Font

GREEN='FFD5F5E3'; RED='FFFADBD8'; NOFILL=PatternFill(fill_type=None)
GST = 1.18

wb = openpyxl.load_workbook('PG_Commercial_Comparison_July2026.xlsx')
cbp = wb['Cost by Payment Mode']

# Apply precise values computed and verified against raw transaction data
cbp.cell(row=7,  column=6).value = 1752486.3708048216   # Credit Card - PayU
cbp.cell(row=7,  column=7).value = 1499111.6335992983   # Credit Card - EaseBuzz
cbp.cell(row=7,  column=8).value = 1767063.4671040217   # Credit Card - Razorpay

cbp.cell(row=11, column=6).value = 32275.754296999996   # Debit Card - PayU
cbp.cell(row=11, column=7).value = 32958.77735735999    # Debit Card - EaseBuzz
cbp.cell(row=11, column=8).value = 38554.19697039999    # Debit Card - Razorpay

cbp.cell(row=12, column=6).value = 26196.985760199997   # Net Banking - PayU
cbp.cell(row=12, column=7).value = 32687.653303899995   # Net Banking - EaseBuzz
cbp.cell(row=12, column=8).value = 30908.278186199997   # Net Banking - Razorpay

# Recolor the three affected rows (cheapest/costliest across the 4 PGs for that row)
def numeric(v):
    return v if isinstance(v,(int,float)) else None
def recolor_row(ws, row, cols):
    for c in cols:
        ws.cell(row=row,column=c).fill = NOFILL
    vals = {c: numeric(ws.cell(row=row,column=c).value) for c in cols}
    nums = {c:v for c,v in vals.items() if v is not None}
    if len(nums) < 2: return
    mn = min(nums.values()); mx = max(nums.values())
    if mn == mx: return
    for c,v in nums.items():
        if v == mn: ws.cell(row=row,column=c).fill = PatternFill('solid', fgColor=GREEN)
        elif v == mx: ws.cell(row=row,column=c).fill = PatternFill('solid', fgColor=RED)

for r in [7,11,12]:
    recolor_row(cbp, r, [5,6,7,8])

# Recompute TOTAL row 19 and effective rate row 21
for c in [5,6,7,8]:
    total = sum(cbp.cell(row=r,column=c).value for r in range(6,19))
    cbp.cell(row=19,column=c).value = total
    cbp.cell(row=21,column=c).value = total / cbp.cell(row=19,column=3).value

totals = {c: cbp.cell(row=19,column=c).value for c in [5,6,7,8]}
print("New TOTAL row:", totals)

# ---- Propagate to Summary Dashboard ----
sd = wb['Summary Dashboard']
pg_rows = {'Cashfree (Actual)':11, 'PayU':12, 'EaseBuzz':13, 'Razorpay':14}
col_map = {11:5, 12:6, 13:7, 14:8}  # Summary Dashboard row -> Cost by Payment Mode col
total_gmv = cbp.cell(row=19,column=3).value
cf_total = totals[5]
for r, cbp_col in col_map.items():
    mdr_cost = sd.cell(row=r, column=2).value  # keep proportion: recompute MDR/GST split proportionally
    total_cost = totals[cbp_col]
    old_total = sd.cell(row=r, column=4).value
    # split total into MDR (pre-GST) and GST(18%) consistently: MDR = total/1.18, GST = total - MDR
    new_mdr = total_cost / GST
    new_gst = total_cost - new_mdr
    sd.cell(row=r, column=2).value = new_mdr
    sd.cell(row=r, column=3).value = new_gst
    sd.cell(row=r, column=4).value = total_cost
    sd.cell(row=r, column=5).value = total_cost / total_gmv
    if r != 11:
        sd.cell(row=r, column=6).value = total_cost - cf_total

# recolor Summary Dashboard comparison rows by total cost (col 4)
rows4 = [11,12,13,14]
tot4 = {r: sd.cell(row=r,column=4).value for r in rows4}
mn_r = min(tot4, key=lambda r: tot4[r]); mx_r = max(tot4, key=lambda r: tot4[r])
neutral_cycle = {11:'FFFFFFFF',12:'FFEBF5FB',13:'FFFFFFFF',14:'FFEBF5FB'}
for r in rows4:
    if r==mn_r: fill=PatternFill('solid', fgColor=GREEN)
    elif r==mx_r: fill=PatternFill('solid', fgColor=RED)
    else: fill=PatternFill('solid', fgColor=neutral_cycle[r])
    for c in range(1,7):
        sd.cell(row=r,column=c).fill = fill

# Update "Cheapest PG" stat box (E4:E6) and "Saves vs Cashfree" text
cheapest_name = sd.cell(row=mn_r, column=1).value.replace(' (Actual)','')
savings_vs_cf = cf_total - tot4[mn_r]
sd.cell(row=5, column=5).value = cheapest_name
sd.cell(row=6, column=5).value = f"Saves ₹{savings_vs_cf:,.0f} vs Cashfree"
print("Cheapest PG:", cheapest_name, "savings:", savings_vs_cf)
print("Most expensive:", sd.cell(row=mx_r,column=1).value, "excess vs CF:", tot4[mx_r]-cf_total)

wb.save('PG_Commercial_Comparison_July2026.xlsx')
