import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import copy

GREEN='FFD5F5E3'; RED='FFFADBD8'; NOFILL=PatternFill(fill_type=None)
PCT='0.00%'

wb = openpyxl.load_workbook('PG_Commercial_Comparison_July2026.xlsx')

# ---- 1. Fix Instrument Rate Matrix Yes Bank row (row 30): PayU & Razorpay were stale April rates ----
ws_irm = wb['Instrument Rate Matrix']
ws_irm.cell(row=30, column=4).value = 0.01   # PayU -> Others rate (not separately quoted)
ws_irm.cell(row=30, column=6).value = 0.012  # Razorpay -> Others rate (not separately quoted)
ws_irm.cell(row=30, column=4).font = Font(italic=True, color='FF666666')
ws_irm.cell(row=30, column=6).font = Font(italic=True, color='FF666666')
ws_irm.cell(row=30, column=8).value = 'PayU/RPY: Yes not separately quoted, Others rate applied'

def numeric(v):
    return v if isinstance(v,(int,float)) else None
def recolor_row(ws, row, cols, cheapest_col=None):
    for c in cols:
        ws.cell(row=row,column=c).fill = NOFILL
    vals = {c: numeric(ws.cell(row=row,column=c).value) for c in cols}
    nums = {c:v for c,v in vals.items() if v is not None}
    if len(nums) < 2: return None
    mn = min(nums.values()); mx = max(nums.values())
    mn_cols=[c for c,v in nums.items() if v==mn]
    for c,v in nums.items():
        if v==mn: ws.cell(row=row,column=c).fill = PatternFill('solid', fgColor=GREEN)
        elif v==mx: ws.cell(row=row,column=c).fill = PatternFill('solid', fgColor=RED)
    return mn_cols

PG_NAME_IRM = {3:'Cashfree',4:'PayU',5:'EaseBuzz',6:'Razorpay'}
mn_cols = recolor_row(ws_irm, 30, [3,4,5,6])
if mn_cols:
    ws_irm.cell(row=30, column=7).value = '/'.join(PG_NAME_IRM[c] for c in mn_cols)

# ---- 2. Rebuild Rate Card Reference Net Banking section (rows 22-25) into bank-level rows ----
ws = wb['Rate Card Reference']
ws.insert_rows(23, amount=4)  # now NET BANKING section needs 7 data rows (23-29) instead of 3 (23-25 before)

# template style from an existing data row (use old row 27, now shifted to row 31, alternating banding source: row7/row8 style for white/lightblue)
white_fill = PatternFill('solid', fgColor='FFFFFFFF')
blue_fill = PatternFill('solid', fgColor='FFEBF5FB')
base_font = copy.copy(ws.cell(row=9, column=1).font)  # a normal data-row font
align = Alignment(horizontal='left')

bank_rows = [
    ('   HDFC Bank',   0.014, 0.0152, 0.014,   0.016,  None),
    ('   ICICI Bank',  0.014, 0.0152, 0.014,   0.016,  None),
    ('   Axis Bank',   0.014, 0.01,   0.014,   0.012,  None),
    ('   SBI',         0.014, 0.01,   0.014,   0.012,  None),
    ('   Yes Bank',    0.014, 0.01,   0.0105,  0.012,  'PayU/RPY: not separately quoted, Others rate applied'),
    ('   Kotak Mahindra', 0.0105, 0.0135, 0.0105, 0.0145, None),
    ('   All Other Banks',0.0105, 0.01,   0.0105, 0.012,  None),
]
for i, (label, cf, payu, eb, rpy, note) in enumerate(bank_rows):
    r = 23 + i
    fill = white_fill if i % 2 == 0 else blue_fill
    ws.cell(row=r, column=1, value=label).fill = fill
    ws.cell(row=r, column=1).font = base_font
    for c, val in zip([2,3,4,5], [cf,payu,eb,rpy]):
        cell = ws.cell(row=r, column=c, value=val)
        cell.number_format = PCT
        cell.fill = fill
    notecell = ws.cell(row=r, column=6, value=note)
    notecell.fill = fill
    if note:
        notecell.font = Font(italic=True, color='FF888888')

for r in range(23, 30):
    recolor_row(ws, r, [2,3,4,5])

wb.save('PG_Commercial_Comparison_July2026.xlsx')
print("Rate Card Reference Net Banking section rebuilt: rows 23-29")
print("New sheet row count:", ws.max_row)
