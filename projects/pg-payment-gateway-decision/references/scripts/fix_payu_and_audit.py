import openpyxl
from openpyxl.styles import PatternFill, Font

GREEN='FFD5F5E3'; RED='FFFADBD8'; NOFILL=PatternFill(fill_type=None)
PCT='0.00%'

wb = openpyxl.load_workbook('PG_Commercial_Comparison_July2026.xlsx')

# ============ 1. Fix PayU uniform 1.75% (except Corporate & CC EMI) ============
ws_rcr = wb['Rate Card Reference']
ws_irm = wb['Instrument Rate Matrix']

# Rate Card Reference: PayU col=3. Diners row9, AMEX row10, International row11
for r in [9,10,11]:
    cell = ws_rcr.cell(row=r, column=3)
    cell.value = 0.0175
    cell.number_format = PCT
    cell.font = Font(italic=False)  # was carried-forward italic; now a real quote (covered by PayU's blanket statement)

# Instrument Rate Matrix: PayU col=4. Diners row16, AMEX row17, International row18
for r in [16,17,18]:
    cell = ws_irm.cell(row=r, column=4)
    cell.value = 0.0175
    cell.number_format = PCT
    cell.font = Font(italic=False)

print("PayU Diners/AMEX/International set to 1.75% (uniform blended rate per PayU's quote; Corporate Cards & CC EMI remain separately quoted)")

# ============ 2. Fix number formats stuck on 'General' ============
general_fixes_rcr = [(8,3),(12,3),(18,5),(20,3)]
for r,c in general_fixes_rcr:
    ws_rcr.cell(row=r,column=c).number_format = PCT

general_fixes_irm = [(7,6),(8,6),(10,4),(15,4),(19,4),(27,5)]
for r,c in general_fixes_irm:
    ws_irm.cell(row=r,column=c).number_format = PCT

print("Number formats normalized to 0.00% for all rate cells")

# ============ 3. Recolor affected rows + update "Cheapest" column in IRM ============
def numeric(v):
    return v if isinstance(v,(int,float)) else None

def recolor_row(ws, row, cols):
    for c in cols:
        ws.cell(row=row,column=c).fill = NOFILL
    vals = {c: numeric(ws.cell(row=row,column=c).value) for c in cols}
    nums = {c:v for c,v in vals.items() if v is not None}
    if len(nums) < 2:
        return None,None
    mn = min(nums.values()); mx = max(nums.values())
    if mn == mx:
        return None,None
    mn_cols=[c for c,v in nums.items() if v==mn]
    mx_cols=[c for c,v in nums.items() if v==mx]
    for c in mn_cols:
        ws.cell(row=row,column=c).fill = PatternFill('solid', fgColor=GREEN)
    for c in mx_cols:
        ws.cell(row=row,column=c).fill = PatternFill('solid', fgColor=RED)
    return mn_cols, mx_cols

PG_NAME_RCR = {2:'Cashfree',3:'PayU',4:'EaseBuzz',5:'Razorpay'}
for r in [8,9,10,11,12,18,20]:
    recolor_row(ws_rcr, r, [2,3,4,5])

PG_NAME_IRM = {3:'Cashfree',4:'PayU',5:'EaseBuzz',6:'Razorpay'}
for r in [7,8,9,10,12,15,16,17,18,19,27]:
    mn_cols, mx_cols = recolor_row(ws_irm, r, [3,4,5,6])
    if mn_cols:
        names = '/'.join(PG_NAME_IRM[c] for c in mn_cols)
        ws_irm.cell(row=r, column=7).value = names

wb.save('PG_Commercial_Comparison_July2026.xlsx')
print("Saved.")
